# -*- coding: utf-8 -*-
"""
Created on Tue Nov 12 11:55:14 2019

@author: Katya Gribkova

Generates Excel data files for SITDL single synapse runs.
Run current Initialization section, then run the desired cell section(s) to generate Excel files.
"""
import openpyxl
import xlsxwriter
import os
import inspect
src_file_path = inspect.getfile(lambda: None)
os.path.join(src_file_path)

import math as mth
import random 
import numpy
import matplotlib.pyplot as plt  
from SITDL_Functions import saveLists, readList, generateSignals, generateRecallSpike, shiftSignal, combinedSaveIndeces,\
                                SITDL, plotSITDLOutput, SITDL_TauLoop, generateInitialTaus, \
                                SITDL_MultiSynapseMemory_Learn, SITDL_MultiSynapseMemory_SynElimination, SITDL_MultiSynapseMemory_Recall, \
                                extractSpikeTimes, extractRiseFallTimes, normalize, calculateDerivatives
                              
#**************INITIALIZATION************************************************

# constants used for all subsequent function calls, unless overidden in section:
t=0.01                                                  # integration time step
dts=0.05                                               # tau_glu time step
num_pts=80000000   #80000000                               # number of points
intervals_list = [30, 66, 48, 72, 90, 54]               # determines repeated sequence of inter-spike intervals
I0 = 0.01                                                 # DC current
a = 3.9                                                  # constant, determines contribution of dendritic signal to voltage
b = 0.4                                                 # constant, determines contribution of synaptic current to voltage
[f, f2]  = generateSignals(num_pts, t, intervals_list, 1, 30., 10., False, False)[1:3]       # signal generation

# %%
#*************************** SITDL SINGLE SYNAPSE RUNS and TAULOOP SIMULATIONS **************************************************
# %%
# SITDL Runs, without Stabilization
num_pts=2000000             # number of points
bookname = "SITDLwoutP SingleRun DATA.xlsx" 
tau_den_list = [10, 15, 45, 95]                     #dendritic delay, based on distance between synapses
tau_glu_list = [50, 5, 5, 50]                       #starting synaptic time lag

varnames = ['tt', 'tau_glu', 'f2_i' , 'Vden', 'gGlu', 'gV', 'g', 'gglimit','g_mismatch', 'plasticity']
labels = ["Time (ms)", "Tau Glu", "Glutamate Signal", "Dendritic Signal", "NMDAR Glutamate Gate Conductance", "NMDAR Voltage Gate Conductance", "NMDAR Conductance", "NMDAR Glutamate Gate Conductance Limit","NMDAR Gate Conductance Mismatch", "Plasticity"]
columns = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
saveparameters = [varnames, labels, columns]
saveranges = list(range(0,30000)) + list(range(num_pts-10000,num_pts))
  
for ts in range(len(tau_den_list)):
    #savetimes = combinedSaveIndeces(tau_den_list[ts], f, f2, num_pts, t, 200, saveranges, 500, 5)
    sheetname = "Tau Den = " + str(round(tau_den_list[ts],1)) + ", Tau Glu = " + str(round(tau_glu_list[ts],1))
    data = SITDL(tau_den_list[ts], tau_glu_list[ts], f, f2, num_pts, t, dts, I0, a, b, False, 2, bookname, sheetname, saveparameters)
    #plotSITDLOutput(data,5,"")

# %%
# SITDL Runs, with Stabilization
num_pts=40000000               # number of points
bookname = "SITDLwithP SingleRun DATA 4.xlsx"           
tau_den_list = [10]                                       #dendritic delay, based on distance between synapses
tau_glu_list = [150]                                       #starting synaptic time lag

varnames = ['tt', 'tau_glu', 'f2_i', 'Vden', 'gGlu', 'gV', 'g', 'g_mismatch', 'plasticity']
labels = ["Time (ms)", "Tau Glu", "Glutamate Signal", "Dendritic Signal", "NMDAR Glutamate Gate Conductance", "NMDAR Voltage Gate Conductance", "NMDAR Conductance", "NMDAR Gate Conductance Mismatch", "Plasticity"]
columns = [3, 4, 5, 6, 7, 8, 9, 10, 11]
saveparameters = [varnames, labels, columns]
saveranges = list(range(0,30000)) + list(range(num_pts-10000,num_pts))
savetime = 10000
  
for ts in range(len(tau_den_list)):
    #savetimes = combinedSaveIndeces(tau_den_list[ts], f, f2, num_pts, t, 200, saveranges, 500, 5)
    sheetname = "Tau Den = " + str(round(tau_den_list[ts],1)) + ", Tau Glu = " + str(round(tau_glu_list[ts],1))
    data = SITDL(tau_den_list[ts], tau_glu_list[ts], f, f2, num_pts, t, dts, I0, a, b, True, savetime, bookname, sheetname, saveparameters)
    #plotSITDLOutput(data,5,"")

# %%
# SITDL Runs, without Stabilization
num_pts=80000000             # number of points
bookname = "SITDLwoutP TauLoop DATA.xlsx" 
tau_den_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, \
                18, 19, 20, 21, 22, 23, 24, 25,26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, \
                41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 55, 60, 65, 70, 75, 80, 85, 90, 95, 100]  #dendritic delays, based on distance between synapses
tau_glu_list = [5]                                                                      #starting synaptic time lag

sheetnames = ["Tau Glu = " + s for s in list(map(str,tau_glu_list))]
varnames = ['tt'] + len(tau_den_list)*['tau_glu']
labels = ["Time (ms)"] + list(map(str,tau_den_list))
columns = list(range(3, 4 + len(tau_den_list)))
saveparameters_loop = [varnames, labels, columns]

SITDL_TauLoop(tau_den_list, tau_glu_list, f, f2, num_pts, t, dts, I0, a, b, False, 100000, bookname, sheetnames, saveparameters_loop)

# %%
# SITDL Runs, with Stabilization
num_pts=80000000             # number of points
bookname = "SITDLwithP TauLoop DATA.xlsx"
tau_den_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, \
                18, 19, 20, 21, 22, 23, 24, 25,26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, \
                41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 55, 60, 65, 70, 75, 80, 85, 90, 95, 100]  #dendritic delays, based on distance between synapses
tau_glu_list = [5, 50]                                                                 #starting synaptic time lag

sheetnames = ["Tau Glu = " + s for s in list(map(str,tau_glu_list))]
varnames = ['tt'] + len(tau_den_list)*['tau_glu']
labels = ["Time (ms)"] + list(map(str,tau_den_list))
columns = list(range(3, 4 + len(tau_den_list)))
saveparameters_loop = [varnames, labels, columns]

SITDL_TauLoop(tau_den_list, tau_glu_list, f, f2, num_pts, t, dts, I0, a, b, True, 100000, bookname, sheetnames, saveparameters_loop)

# %%
#*************************** Supplementary Data **************************************************
# %%
# For comparing (gGlu - gV) and (gL - gGlu) evolution
num_pts=18000000             # number of points
bookname = "SITDL (gL-gGlu) vs (gGlu-gV) DATA.xlsx"
sheetname = 'Examples'
varnames = ['tt', 'tau_glu', 'gGlu', 'gV', 'gglimit']
labels = ["Time (ms)", "Tau Glu", "NMDAR Glutamate Gate Conductance", "NMDAR Voltage Gate Conductance", "NMDAR Glutamate Gate Conductance Limit"]
columns = [3, 4, 5, 6, 7]
saveparameters = [varnames, labels, columns]
saveranges = list(range(0,30000,2)) + list(range(950000,980000,2)) + list(range(17300000,17900000,2))
data = SITDL(10, 50, f, f2, num_pts, t, dts, I0, a, b, False, saveranges, bookname, sheetname, saveparameters)

sheetname = 'Phase Plot'
data = SITDL(10, 50, f, f2, num_pts, t, dts, I0, a, b, False, 500, bookname, sheetname, saveparameters)

# %%
num_pts=500000 
bookname = "SITDL SignalGeneration DATA.xlsx"
varnames = ['tt', 'f_i', 'f2_i']
labels = ["Time (ms)", "SN", "SL"]
columns = [3, 4, 5]
saveparameters = [varnames, labels, columns]

intervals_list = [20, 14, 22, 30]
[tg, f, f2]  = generateSignals(num_pts, t, intervals_list, 1, 30., 10., True, False)       # signal generation
sheetname = 'Signals-SITDL_MultiSyn'
data = SITDL(0, 10, f, f2, num_pts, t, dts, I0, a, b, False, 1, bookname, sheetname, saveparameters)

intervals_list = [30, 66, 48, 72, 90, 54] 
[tg, f, f2]  = generateSignals(num_pts, t, intervals_list, 1, 30., 10., False, False)       # signal generation
sheetname = 'Signals-SITDL_SingleSyn'
data = SITDL(0, 10, f, f2, num_pts, t, dts, I0, a, b, False, 1, bookname, sheetname, saveparameters)

# %%
num_pts=50000 #2000000                                # number of points
[tg, f, f2]  = generateRecallSpike(num_pts, t, 0, 1, 30., 10.)      # signal generation
bookname = "SITDL TauSyn vs TauGlu.xlsx"
tau_den = 0                    #dendritic delay, based on distance between synapses
tau_glu_list = [*range(1,10001)]                       #fixed synaptic time constant

varnames = ['tt', 'tau_glu', 'f2_i', 'Vden', 'gGlu', 'gV', 'g', 'g_mismatch', 'plasticity']
labels = ["Time (ms)", "Tau Glu", "Glutamate Signal", "Dendritic Signal", "NMDAR Glutamate Gate Conductance", "NMDAR Voltage Gate Conductance", "NMDAR Conductance", "NMDAR Gate Conductance Mismatch", "Plasticity"]
columns = [3, 4, 5, 6, 7, 8, 9, 10, 11]
saveparameters = [varnames, labels, columns]

risetimes = len(tau_glu_list) * [0]
peaktimes = len(tau_glu_list) * [0]
  
for ts in range(len(tau_glu_list)):
    sheetname = ''#"Tau Den = " + str(round(tau_den,1)) + ", Tau Glu = " + str(round(tau_glu_list[ts],1))
    data = SITDL(tau_den, tau_glu_list[ts], f, f2, num_pts, t, dts, I0, a, b, -1, 1, bookname, sheetname, saveparameters) #note no changes in tau_glu
    tg = data[0]
    gg = data[4]
    threshold = 1 + 1 / (1000 * tau_glu_list[ts])
    pt = extractSpikeTimes(tg, gg, 10)[0]
    peaktimes[ts] = pt[0]
    risetimes[ts] = pt[0] - 0.9 #rise starts at 0.9 ms
saveLists([tau_glu_list, peaktimes, risetimes], 1, bookname, 'tau_glu to Rise Time', [['tau_glu_list', 'peaktimes', 'risetimes'], ['tau_glu', 'Peak-times (ms)','Rise-times (ms)'], [2,3,4]])
