# -*- coding: utf-8 -*-
"""
Created on Tue Nov 12 11:55:14 2019

@author: Ekaterina Gribkova

Functions for all SITDL Simulations:
"""
import openpyxl
import xlsxwriter
import os
import math as mth
import random 
import numpy 
import matplotlib.pyplot as plt  


# **************************Data Handling Functions for Excel*******************
def setheadingsExcelSheet(sheet, saveparameters):
    labels = saveparameters[1]
    columns = saveparameters[2]
    for i in range(len(labels)):
        if columns[i] > 0:
            sheet.cell(row = 1, column = columns[i]).value = labels[i]

def writetoExcelSheet(sheet, nrow, saveparameters, data):
    labels = saveparameters[1]
    columns = saveparameters[2]
    for i in range(len(labels)):
        if columns[i] > 0:
            sheet.cell(row = nrow, column = columns[i]).value = data[i]
                       
def saveLists(datalists, savetime, bookname, sheetname, saveparameters):   ############ DO NOT USE FOR LARGE LISTS! TOO SLOW
    n = 0
    if bookname != '' and sheetname != '':
        [book, sheet] = openExcelDoc(bookname, sheetname)
        for i in range(len(datalists[0])):
            if (i % savetime == 0):
                n = n + 1
                if n == 1:
                    setheadingsExcelSheet(sheet, saveparameters)
                data = (numpy.array(datalists)[:,i]).tolist()
                writetoExcelSheet(sheet, n+1, saveparameters, data)
        book.save(bookname)
        book.close()                     

def openExcelDoc(bookname, sheetname):
    if os.path.isfile(bookname): book = openpyxl.load_workbook(bookname)
    else: book = openpyxl.Workbook()
    if sheetname not in book.sheetnames: book.create_sheet(sheetname)
    sheet = book[sheetname]
    return book, sheet
    
def findExcelSheet(bookname, sheetname):
    book = False
    sheet = False
    if os.path.isfile(bookname): 
        book = openpyxl.load_workbook(bookname)
        if sheetname in book.sheetnames: sheet = book[sheetname]
    return book, sheet  
    
def readList(columnname, bookname, sheetname):
    [book, sheet] = findExcelSheet(bookname, sheetname)
    datalist = []
    if sheet is not False:
        max_col = sheet.max_column
        max_row = sheet.max_row
        for i in range(1, max_col+1):
            if columnname == sheet.cell(row = 1, column = i).value:
                for j in range(2,max_row+1):
                    datapoint = sheet.cell(row = j, column = i).value
                    if datapoint is not None: datalist.append(datapoint)
    return datalist
    
# %%
#**************SIGNAL (SPIKE PATTERN) GENERATION**************************
#spike_width: gauss spike maximum duration, sig_gauss: gauss spike dispersion, al: glutamate spike decay  
def generateSignals(num_pts, time_step, intervals_list, spike_width, sig_gauss, al, Vsparse, Gsparse):
    #....................Generate (voltage) signal with gauss-like spikes...............
    tg = [0]*num_pts
    f =[0]*num_pts
    tt = 0.                       #start time
    t0 = spike_width              #gauss spike maximum duration       
    dt = t0
    t0v = [t0 - 1/al]             # peak times, shifted by 1/al
    
    fsp =[0]*num_pts              # for sparse signals (only 1st spike of each period)
    t0sp = [t0]
    t0vsp = [t0 - 1/al]
    
    
    count = -1
    for j in range(num_pts):
        tt = tt + time_step
        tg[j] = tt
        if tt >= t0 + dt:
           count = count + 1
           rnd = intervals_list[count % len(intervals_list)]#rnd = 2.+ 3*random.random()
           t0 = tt + dt*rnd
           t0v.append(t0 - 1/al)
           if count % len(intervals_list) == len(intervals_list) - 1: t0sp.append(t0)
        f[j]= mth.exp( -((tt - t0)**2*sig_gauss))

    t0 = spike_width
    tt = 0    
    count = -1
    if Vsparse or Gsparse:          # for sparse signal generation
        for j in range(num_pts):
            tt = tt + time_step
            tg[j] = tt
            if tt >= t0 + dt and count < len(t0sp)-1:
               count = count + 1
               t0 = t0sp[count]
               t0vsp.append(t0 - 1/al)
            fsp[j]= mth.exp( -((tt - t0)**2*sig_gauss))
        if Vsparse: f = fsp
        if Gsparse: t0v = t0vsp
    #....................Generate (glutamate) signal with log-normal-like spikes...............
    f2 =[0]*num_pts
    tt= 0               #start time
    t0 = 0
    count = 0
    
    for j in range(num_pts):
        tt = tt + time_step
        if count == 0: t0 = tt
        if tt >= t0v[count] and count < len(t0v)-1:
            count = count + 1
            t0 = tt
        f2[j]= ((tt-t0))*(mth.exp(-((tt-t0))*al))#Isig


    #....................Normalization.........................................
    fmax=max(f)
    f2max = max(f2)    
    for j in range(num_pts):
        f[j]=f[j]/fmax
        f2[j]=f2[j]/f2max        

    return tg,f,f2
 
def generateRecallSpike(num_pts, time_step, time, spike_width, sig_gauss, al):
    #....................Generate (voltage) signal with gauss-like spike...............
    tg = [0]*num_pts
    f =[0]*num_pts
    tt = 0.                       #start time
    t0 = spike_width              #gauss spike maximum duration       
    dt = t0
    t02 = t0 - 1/al             # peak times, shifted by 1/al

    for j in range(num_pts):
        tt = tt + time_step
        tg[j] = tt            
        f[j]= mth.exp( -(((tt - t0)**2)*sig_gauss))
        
    #....................Generate (glutamate) signal with log-normal-like spike...............
    f2 =[0]*num_pts
    tt= 0               #start time
    t0 = 0    
    for j in range(num_pts):
        tt = tt + time_step
        if tt < t02:
            t0 = tt
        f2[j]= ((tt-t0))*(mth.exp(-((tt-t0))*al))#Isig
        
    #....................Normalization.........................................
    fmax=max(f)
    f2max = max(f2)    
    for j in range(num_pts):
        f[j]=f[j]/fmax
        f2[j]=f2[j]/f2max        

    return tg,f,f2

# %%    
#**********************************************************************************
#             SITDL MODEL SINGLE TAU
#**********************************************************************************
def SITDL(tau_den, tau_glu, f, f2, num_pts, t, dts, I0, a, b, stabilization, savetime, bookname, sheetname, saveparameters):
    print("Tau Den = " + str(round(tau_den,1)) + ", Tau Glu = " + str(round(tau_glu,1)) + ", Stabilization = " + str(stabilization))
    #....................Generate (glutamate) signal with log-normal-like spikes...............    
    g_mismatch = 1.             # mismatch/difference between NMDAR glutamate gate and voltage gate conductances 
    ss1 = 0.                    # accumulates when NMDAR conductance mismatch is low and g is high
    plasticity = 1.             # determines when synaptic tau glu can change, depends on NMDAR conductance mismatch and stability sum
    gmemory = 0.                 # for memorizing last highest g value
    #..................................................................
    #I0=0.04                     # DC current
    Isig=0.01                   # Signal (of randomly generated spikes) for dendritic and synaptic signals
    Isyn=0.                     # Synaptic current output: Isyn=g*(Vden-Vthresh) 
    Vden=0.                     # denritic voltage 
    Vrest=0.                    # Rest potential
    Vthresh=0.                  # Synaptic threshold
    tau_rest=1.                 #
    #.........NMDAR voltage gates........................................
    gV=0.                       # Voltage gates conductance for NMDARs
    #.........NMDAR glutamate gates......................................
    gGlu=0.                     # Glu gates conductance for NMDARs 
    gglimit=0.                    # initial Glu gates max conductance
    tau_glu_min = 5.            # minimum value for tau glu, which corresponds to minimum rise time of around 7 ms

    #.........Graphing/Saving......................................
    tg = []; sg = []; gg = []; vg = []; cgg = []; cvg = [];  cg = []; cmg = []; pg = []; mgg = []; vfull = [];
    if bookname != '' and sheetname != '': [book, sheet] = openExcelDoc(bookname, sheetname)
    count = 0
    #******************MAIN LOOP****************
    tt = 0.
    n = 0
    md = int(tau_den/t)                                                 # denritic delay in points (point index for shift)
    for i in range(num_pts):
          tt=tt+t
          II=0.
          if i >= md : II = I0 + f[i-md]                                #delayed dendritic signal
    #***************Model**************************************** 
          gV = 1/(1.+ mth.exp(-8.*Vden + 5))                            # conductance through NMDAR voltage gates
          g_mismatch = (gGlu - gV)                                      # NMDAR conductance mismatch: difference between NMDAR glutamate gate and voltage gate conductances 
          dtau_glu = dts*g_mismatch*(gglimit - gGlu)                      # change in tau_glu
          tau_glu = tau_glu + plasticity*dtau_glu                       # time constant with which gGlu rises to its limit
          if tau_glu <= tau_glu_min : tau_glu = tau_glu_min             # min limit on synaptic tau                       
          
          g = gGlu * gV / (gGlu + gV)                                   # combined (in-series) NMDAR conductance from voltage and glutamate gate conductances
          Isyn = g*(Vden - Vthresh)                                     # current trhough NMDAR channels, based on conductance and voltage
          Vden = Vden + t*((-Vden + Vrest)/tau_rest + a*II + b*Isyn)    # Dendritic voltage, with contribution from synaptic current
                 
          gglimit = 0.999 * gglimit + 0.065 * f2[i]                         # limit of gGlu, dependent on glutamate signal
          gGlu = gglimit + (gGlu - gglimit)*mth.exp(-t/tau_glu)
                                                 
          if tt > 0.999 * t * num_pts: gmemory = gmemory + g / (0.001 * num_pts)  # determining last g values attained for synaptic elimination
          if ss1 < 2000: ss1 = ss1 + 0.05*(dts/4 - abs(dtau_glu))*g        # slowly accumulates when dtau_glu is close to zero, and g is non-zero; condition on ss1 is just to prevent math overflow error  
          if stabilization == True:
              plasticity = (1./(1. + mth.exp(0.3*ss1 - 70)))            # determines when synaptic tau can change, depends on stability sum variable
              #print ("SS1 = " + str(round(ss1,1)) + ", P = " + str(round(plasticity,10)))
          else:
              if stabilization == False: 
                  plasticity = 1                 # always plastic in this case
              if stabilization == -1: 
                  plasticity = 0                 # never changing in this case. Note, ss1 is still updated 
                  vfull.append(Vden)             # for Recall
          #     ......printed output..........
          if i % int(num_pts/50) == 0: print('%',end='')   #a "progress bar"
          if i == (num_pts - 1): print ("\n      gmemory = " + str(gmemory) + ", TauGlu = " + str(tau_glu) + ", SS1 = " + str(round(ss1,1)) + ", P = " + str(round(plasticity,10)))
          #     ......for graphing/saving..........
          savepoint = False      #saving sparse data
          if isinstance(savetime,list) and count < len(savetime) and i == savetime[count]: 
              savepoint = True 
              count = count + 1
          if isinstance(savetime,int) and i % savetime == 0: savepoint = True 
          if savepoint:
              f_i = f[i]
              f2_i = f2[i]
              tg.append(tt); sg.append(tau_glu); gg.append(f2_i); vg.append(Vden);          ###########
              cgg.append(gGlu); cvg.append(gV); cg.append(g); cmg.append(g_mismatch);
              pg.append((plasticity)); mgg.append(gmemory);
              if bookname != '' and sheetname != '':            #saving to excel workbook
                  n = n + 1
                  if n == 1: setheadingsExcelSheet(sheet, saveparameters)
                  data = len(saveparameters[0])*[0]
                  for s in range(len(data)): data[s] = locals() [saveparameters[0][s]]
                  writetoExcelSheet(sheet, n+1, saveparameters, data)
    if bookname != '' and sheetname != '':
        book.save(bookname)
        book.close()
    return tg, sg, gg, vg, cgg, cvg, cg, cmg, pg, gmemory, vfull

# %%
#same SITDL function, but saves only tau_glu data (less memory usage)
def SITDL_SaveTauOnly(tau_den, tau_glu, f, f2, num_pts, t, dts, I0, a, b, stabilization, savetime, bookname, sheetname, saveparameters):
    print("Tau Den = " + str(round(tau_den,1)) + ", Tau Glu = " + str(round(tau_glu,1)) + ", Stabilization = " + str(stabilization))
    #....................Generate (glutamate) signal with log-normal-like spikes...............    
    g_mismatch = 1.             # mismatch/difference between NMDAR glutamate gate and voltage gate conductances 
    ss1 = 0.                    # accumulates when NMDAR conductance mismatch is low and g is high
    plasticity = 1.             # determines when synaptic tau can change, depends on NMDAR conductance mismatch and stability sum
    gmemory = 0.                 # for memorizing last highest g value
    #..................................................................
    #I0=0.04                     # DC current
    Isig=0.01                   # Signal (of randomly generated spikes) for dendritic and synaptic signals
    Isyn=0.                     # Synaptic current output: Isyn=g*(Vden-Vthresh) 
    Vden=0.                     # denritic voltage 
    Vrest=0.                    # Rest potential
    Vthresh=0.                  # Synaptic threshold
    tau_rest=1.                 #
    #.........NMDAR voltage gates........................................
    gV=0.                       # Voltage gates conductance for NMDARs
    #.........NMDAR glutamate gates......................................
    gGlu=0.                     # Glu gates conductance for NMDARs 
    gglimit=0.                    # initial Glu gates conductance limit
    tau_glu_min = 0.            # minimum value for tau syn, which corresponds to minimum rise time of around 7 ms
    #.........Graphing/Saving......................................
    if bookname != '' and sheetname != '': [book, sheet] = openExcelDoc(bookname, sheetname)
    count = 0
    #******************MAIN LOOP****************
    tt = 0.
    n = 0
    md = int(tau_den/t)                                                 # denritic delay in points (point index for shift)
    for i in range(num_pts):
          tt=tt+t
          II=0.
          if i >= md : II = I0 + f[i-md]                                #delayed dendritic signal
    #***************Model**************************************** 
          gV = 1/(1.+ mth.exp(-8.*Vden + 5))                            # conductance through NMDAR voltage gates
          g_mismatch = (gGlu - gV)                                      # NMDAR conductance mismatch: difference between NMDAR glutamate gate and voltage gate conductances 
          dtau_glu = dts*g_mismatch*(gglimit - gGlu)                      # change in tau_glu
          tau_glu = tau_glu + plasticity*dtau_glu                       # time constant with which gGlu rises to its limit
          if tau_glu <= tau_glu_min : tau_glu = tau_glu_min             # min limit on synaptic tau                       
          
          g = gGlu * gV / (gGlu + gV)                                   # combined (in-series) NMDAR conductance from voltage and glutamate gate conductances
          Isyn = g*(Vden - Vthresh)                                     # current trhough NMDAR channels, based on conductance and voltage
          Vden = Vden + t*((-Vden + Vrest)/tau_rest + a*II + b*Isyn)    # Dendritic voltage, with contribution from synaptic current
                 
#          gglimit=1/(1.+mth.exp(-10.*f2[i] +3.))                          # limit of gGlu, dependent on glutamate signal
          gglimit = 0.999 * gglimit + 0.065 * f2[i]                         # limit of gGlu, dependent on glutamate signal
          gGlu = gglimit + (gGlu - gglimit)*mth.exp(-t/tau_glu)
                                                 
          if tt > 0.999 * t * num_pts: gmemory = gmemory + g / (0.001 * num_pts)  # determining last g values attained for synaptic elimination
          if ss1 < 2000: ss1 = ss1 + 0.05*(dts/4 - abs(dtau_glu))*g        # slowly accumulates when dtau_glu is close to zero, and g is non-zero; condition on ss1 is just to prevent math overflow error  
          if stabilization == True:
              plasticity = (1./(1. + mth.exp(0.3*ss1 - 70)))            # determines when synaptic tau can change, depends on stability sum variable
              #print ("SS1 = " + str(round(ss1,1)) + ", P = " + str(round(plasticity,10)))
          else:
              if stabilization == False: 
                  plasticity = 1                 # always plastic in this case
              if stabilization == -1: 
                  plasticity = 0                 # never changing in this case. Note, ss1 is still updated 
          #     ......printed output..........
          if i % int(num_pts/50) == 0: print('%',end='')   #a "progress bar"
          if i == (num_pts - 1): print ("\n      gmemory = " + str(gmemory) + ", Tau Glu = " + str(tau_glu) + ", SS1 = " + str(round(ss1,1)) + ", P = " + str(round(plasticity,10)))
          #     ......for graphing/saving..........
          savepoint = False      #saving sparse data
          if isinstance(savetime,list) and count < len(savetime) and i == savetime[count]: 
              savepoint = True 
              count = count + 1
          if isinstance(savetime,int) and i % savetime == 0: savepoint = True 
          if savepoint:
              if bookname != '' and sheetname != '':            #saving to excel workbook
                  n = n + 1
                  if n == 1: setheadingsExcelSheet(sheet, saveparameters)
                  data = len(saveparameters[0])*[0]
                  for s in range(len(data)): data[s] = locals() [saveparameters[0][s]]
                  writetoExcelSheet(sheet, n+1, saveparameters, data)
    if bookname != '' and sheetname != '':
        book.save(bookname)
        book.close()
    return tau_glu, plasticity, gmemory   
# %%
#same SITDL function, but returns minimal data (last tau_glu, gmemory, and plasticity values), for much less memory usage
def SITDL_nosave(tau_den, tau_glu, f, f2, num_pts, t, dts, I0, a, b, stabilization): 
    print("Tau Den = " + str(round(tau_den,1)) + ", Tau Glu = " + str(round(tau_glu,1)) + ", Stabilization = " + str(stabilization))
    #....................Generate (glutamate) signal with log-normal-like spikes...............    
    g_mismatch = 1.             # mismatch/difference between NMDAR glutamate gate and voltage gate conductances 
    ss1 = 0.                    # accumulates when NMDAR conductance mismatch is low and g is high
    plasticity = 1.             # determines when synaptic tau can change, depends on NMDAR conductance mismatch and stability sum
    gmemory = 0.                 # for memorizing last highest g value
    #..................................................................
    #I0=0.04                     # DC current
    Isig=0.01                   # Signal (of randomly generated spikes) for dendritic and synaptic signals
    Isyn=0.                     # Synaptic current output: Isyn=g*(Vden-Vthresh) 
    Vden=0.                     # denritic voltage 
    Vrest=0.                    # Rest potential
    Vthresh=0.                  # Synaptic threshold
    tau_rest=1.                 #
    #.........NMDAR voltage gates........................................
    gV=0.                       # Voltage gates conductance for NMDARs
    #.........NMDAR glutamate gates......................................
    gGlu=0.                     # Glu gates conductance for NMDARs 
    gglimit=0.                    # initial Glu gates max conductance
    tau_glu_min = 5.            # minimum value for tau glu, which corresponds to minimum rise time of around 7 ms
    #******************MAIN LOOP****************
    tt = 0.
    n = 0
    md = int(tau_den/t)                                                 # denritic delay in points (point index for shift)
    for i in range(num_pts):
          tt=tt+t
          II=0.
          if i >= md : II = I0 + f[i-md]                                #delayed dendritic signal
    #***************Model**************************************** 
          gV = 1/(1.+ mth.exp(-8.*Vden + 5))                            # conductance through NMDAR voltage gates
          g_mismatch = (gGlu - gV)                                      # NMDAR conductance mismatch: difference between NMDAR glutamate gate and voltage gate conductances 
          dtau_glu = dts*g_mismatch*(gglimit - gGlu)                      # change in tau_glu
          tau_glu = tau_glu + plasticity*dtau_glu                       # time constant with which gGlu rises to its limit
          if tau_glu <= tau_glu_min : tau_glu = tau_glu_min             # min limit on synaptic tau                       
          
          g = gGlu * gV / (gGlu + gV)                                   # combined (in-series) NMDAR conductance from voltage and glutamate gate conductances
          Isyn = g*(Vden - Vthresh)                                     # current trhough NMDAR channels, based on conductance and voltage
          Vden = Vden + t*((-Vden + Vrest)/tau_rest + a*II + b*Isyn)    # Dendritic voltage, with contribution from synaptic current
                 
          gglimit = 0.999 * gglimit + 0.065 * f2[i]                         # limit of gGlu, dependent on glutamate signal
          gGlu = gglimit + (gGlu - gglimit)*mth.exp(-t/tau_glu)
#                                                                       # determining maximum g attained for synaptic elimination
          if tt > 0.999 * t * num_pts: gmemory = gmemory + g / (0.001 * num_pts)
          if ss1 < 2000: ss1 = ss1 + 0.05*(dts/4 - abs(dtau_glu))*g        # slowly accumulates when dtau_glu is close to zero, and g is non-zero; condition on ss1 is just to prevent math overflow error  
          if stabilization == True:
              plasticity = (1./(1. + mth.exp(0.3*ss1 - 70)))            # determines when synaptic tau can change, depends on stability sum variable
          else:
              if stabilization == False: plasticity = 1                 # always plastic in this case
              if stabilization == -1: plasticity = 0                    # never changing in this case. Note, ss1 is still updated  
          #     ......printed output..........
          if  i % int(num_pts/50) == 0: print('%',end='')   #a "progress bar"
          if i == (num_pts - 1): print ("\n      gmemory = " + str(gmemory) + ", TauGlu = " + str(tau_glu) + ", SS1 = " + str(round(ss1,1)) + ", P = " + str(round(plasticity,10)))
    
    return tau_glu, plasticity, gmemory                                 #returns only last points, uses minimal amout of memory

# %%
def plotSITDLOutput(data, start_index, end_index, graphsplit_index, graph_title):
    [tg, sg, gg, vg, cgg, cvg, cg, cmg, pg, mgg, vfull] = data
    graphs = [ vg, gg, cgg, cvg, cg, sg, cmg, pg]
    labels = ["Denritic signal", "Glutamate signal", "gGlu", "gV", "g", "Tau Glu", "Glu and V gate conductance mismatch", "P"]
    colors = ["lightskyblue", "lightgreen", "green", "blue", "red", "black", "red", "violet"]
    linewidths = [3, 3, 2, 2, 2, 1, 1, 1]

    for i in range(len(graphs)):
        if i >= start_index and i<= end_index:
            if i == start_index or i == graphsplit_index: #5
                plt.figure(figsize=(20, 7), dpi=100)
                plt.xlabel("time (ms)")
            plt.plot(tg, graphs[i], color=colors[i], label=labels[i], linewidth=linewidths[i])
            plt.legend(loc = 'upper right', bbox_to_anchor = (0.9, 1.15), ncol = 2)
            plt.title(graph_title)
    plt.show()
 
# %% 
#**************SITDL Simulations for a list of initial tau_den, tau_glu**************************
def SITDL_TauLoop(tau_den_list, tau_glu_list, f, f2, num_pts, t, dts, I0, a, b, stabilization, savetime, bookname, sheetnames, saveparameters_loop):
    for ts in range (len(tau_glu_list)):
        sheetname = sheetnames[ts]
        for td in range (len(tau_den_list)):
            saveparameters = (numpy.array(saveparameters_loop)[:,[0,td+1]]).tolist() #for saving tau glu data from loops onto a single sheet
            saveparameters[2] = list(map(int,saveparameters[2]))
            SITDL_SaveTauOnly(tau_den_list[td], tau_glu_list[ts], f, f2, num_pts, t, dts, I0, a, b, stabilization, savetime, bookname, sheetname, saveparameters)

# %%  
#*****************Small functions for next SITDL MultiSynapseMemory************************************
def generateInitialTaus(num_synapses, tau_den_min, tau_den_max, tau_glu_min, tau_glu_max):
    tau_den_list = [0]*num_synapses
    tau_glu_list = [0]*num_synapses
    for ts in range (num_synapses):
        tau_den_list[ts] = tau_den_min + ts*(tau_den_max - tau_den_min)/(num_synapses-1)
        tau_glu_list[ts] = tau_glu_min + ts*(tau_glu_max - tau_glu_min)/(num_synapses-1)
    return tau_den_list, tau_glu_list    
        
def summateVoltages(tau_den_list, tau_glu_list, stable_syns, t, Vden_array): 
    num_synapses = len(tau_den_list)
    num_ss = len(stable_syns)
    num_pts = len(Vden_array[0])
    Vfinal = [0]*num_pts                                
    Vcontrol = [0]*num_pts                              
    Vsyns = numpy.zeros((num_ss,num_pts))               #first num_syn element of Vden_array for SITDL V
    Vsyns_control = numpy.zeros((num_ss,num_pts))       #last element of Vden_array for non-SITDL control V
    for i in range(num_pts):      
            for j in range (num_ss):
                ts = stable_syns[j]
                tau_glu = tau_glu_list[ts]
                tau_den = tau_den_list[ts]               #dendritic delay, based on distance between synapses
                md = int(tau_den/t)                      # denritic delay in points (point index for shift)
                if i >= md : 
                    Vsyns[j, i] = Vden_array[ts,i-md]
                    Vsyns_control[j, i] = Vden_array[num_synapses - 1,i-md]
    Vfinal = numpy.sum(Vsyns, axis = 0).tolist()
    Vcontrol = numpy.sum(Vsyns_control, axis = 0).tolist()
    return Vsyns, Vsyns_control, Vfinal, Vcontrol
    
def extractSpikeTimes(tg, V, windowlength):
    num_pts = len(tg)
    spiketimes = []
    peakvalues = []
    w = windowlength
    h = int(w/2)
    for i in range(num_pts):
        if i >= w:
            mid = i-h
            peak = True
            j = 0
            while peak and j < h:
                j = j + 1
                if V[mid-j] >= V[mid] or V[mid] <= V[mid+j]:
                    peak = False
            if peak:
                spiketimes.append(tg[mid])
                peakvalues.append(V[mid]) 
    return spiketimes, peakvalues
    
def extractRiseFallTimes(tg, g, windowlength, threshold):
    num_pts = len(tg)
    starttimes = []
    peaktimes = []
    endtimes = []
    w = windowlength
    h = int(w/2)
    for i in range(num_pts):
        if i > 30 and i >= w:
            mid = i-h
            peak = True
            j = 0
            while peak and j < h:
                j = j + 1
                if g[mid-j] >= g[mid] or g[mid] <= g[mid+j]:
                    peak = False
            
            if len(starttimes) <= len(peaktimes): rise = True
            else: rise = False
            
            fall = True
            if len(endtimes) > len(starttimes): endtimes.pop(0)
            
            j = 0
            while (rise or fall) and j < w:
                j = j + 1
                if g[i-j] >= g[i] or threshold * g[i-w] >= g[i]: rise = False
                if g[i-j] <= g[i] or g[i-w] <= threshold * g[i]: fall = False

            if rise: starttimes.append(tg[i-w])
            if peak: peaktimes.append(tg[mid])
            if fall: endtimes.append(tg[i])
    mincount = min(len(starttimes), len(peaktimes),len(endtimes))
    risetimes = [0]*mincount
    falltimes = [0]*mincount
    for i in range(mincount):
        risetimes[i] = peaktimes[i] - starttimes[i]
        falltimes[i] = endtimes[i] - peaktimes[i]
    return peaktimes, risetimes, falltimes
# %%

#**********************************************************************************
#             SITDL FUNCTIONS FOR MULIT-SYNAPSE MEMORY
#**********************************************************************************
# LEARNING PHASE: Takes list of synapses (with initial tau_den, tau_glu) through SITDL simulations
def SITDL_MultiSynapseMemory_Learn(tau_den_list, tau_glu_list, f, f2, num_pts, t, dts, stabilization, savetime, bookname, sheetname, save_to_Excel):
    print("Learning Phase. " + str(len(tau_den_list)) + " synapses, TauD from " + str(min(tau_den_list)) + " to " + str(max(tau_den_list)))
    num_synapses = len(tau_den_list)
    mem_list = [0]*num_synapses
    plasticity_list = [0]*num_synapses
    # Learning and Tau Glu memorization
    for ts in range (num_synapses):
            data = SITDL_nosave(tau_den_list[ts], tau_glu_list[ts], f, f2, num_pts, t, dts, 0.04, 3.9, 0.4, stabilization)
            #plotSITDLOutput(data, 0)
            tau_glu_list[ts] = data[0]
            plasticity_list[ts] = data[1]
            mem_list[ts] = data[2]
    if save_to_Excel: saveLists([tau_den_list, tau_glu_list, plasticity_list, mem_list],1, bookname, sheetname, [[''],["TauD","TauGlu", "Plasticity", "Gmemory"],[3,4,5,6]])
    return tau_glu_list, plasticity_list, mem_list

# SYNAPTIC ELIMINATION: Eliminates synapses with insufficient average NMDAR conductance, based on population average and elimination threshold 
def SITDL_MultiSynapseMemory_SynElimination(tau_den_list, tau_glu_list, mem_list, elim_factor):
    #Synaptic elimination (based on threshold)
    num_synapses = len(tau_den_list)
    stable_syns = []
    mem_avg = sum(mem_list) / len(mem_list)
    ts = 0
    print("Synapse Elimination. Threshold Factor = " +  str(elim_factor) + ", Average g = " + str(round(mem_avg,6)))
    while ts < num_synapses:
        if mem_list[ts] < elim_factor * mem_avg: 
            tau_glu_list[ts] = -1               #synapse marked as eliminated
        else:
            stable_syns.append(ts)
        ts = ts + 1
    return stable_syns, tau_glu_list

# RECALL PHASE: "Recall" of original signal, using coincident glutamate and voltage spike (AMPA) and memorized Tau Syns
def SITDL_MultiSynapseMemory_Recall(tau_glu_initial, stable_syns, tau_den_list, tau_glu_list, mem_list, f, fs, f2, fs2, recall_num_pts, t, dts, savetime_recall, bookname, sheetnames, save_to_Excel):
    print("Recall Phase. " +  str(len(stable_syns)) + " synapses stabilized")
    num_pts = recall_num_pts
    num_synapses = len(tau_den_list)

    tau_glu_list.append(tau_glu_initial)
    tau_den_list.append(0)
    Vden_array = numpy.zeros((num_synapses+1,num_pts))
    for ts in range (num_synapses+1):
            if ts < num_synapses:
                tau_glu=tau_glu_list[ts]                #memorized synaptic time lag
            else:
                tau_glu = tau_glu_initial
            tau_den  = 0
            if tau_glu >= 0:
                data = SITDL(tau_den, tau_glu, fs, fs2, num_pts, t, 0, 0., 2.0, 3.0, -1, savetime_recall, "", "", "")
                Vden_array[ts,:] = data[-1]
                del data

    #Recall: summate shifted synaptic voltages
    sheetname = sheetnames[1]
    [Vsyns, Vsyns_control, Vfinal, Vcontrol] = summateVoltages(tau_den_list, tau_glu_list, stable_syns, t, Vden_array)
    num_ss = len(stable_syns)
    tau_dens_stable = [tau_den_list[index] for index in stable_syns]
    varnames = ['tt'] + num_ss*['Vsyns_ts_i'] + [ 'f2_i', 'Vfinal_i', 'Vcontrol_i'] + num_ss*['Vsynscon_ts_i']
    labels = ["Time (ms)"] + list(map(str,tau_dens_stable)) + ["Original Glutamate Signal", "Signal Recall", "Signal Recall No Shift"] + list(map(str,tau_dens_stable))
    columns = list(range(3, 7 + 2*num_ss))
    saveparameters_loop = [varnames, labels, columns]
    tg = num_pts * [0]
    if save_to_Excel and bookname != '' and sheetname != '':
        [book, sheet] = openExcelDoc(bookname, sheetname)
        for ts in range(num_ss):
            indices = [0,ts+1] + list(range(num_ss + 1,num_ss + 4)) + [num_ss + 4 + ts]
            saveparameters = (numpy.array(saveparameters_loop)[:,indices]).tolist() #for saving synaptic V data from loops onto a single sheet
            saveparameters[2] = list(map(int,saveparameters[2]))
            n = 0
            tt = 0
            for i in range(num_pts):
                tt = tt + t
                tg[i] = tt
                f_i = f[i]
                f2_i = f2[i]
                Vsyns_ts_i = Vsyns[ts,i]
                Vsynscon_ts_i = Vsyns_control[ts, i]
                Vfinal_i  = Vfinal[i]
                Vcontrol_i = Vcontrol[i]
                if (i % savetime_recall == 0) and bookname != '':          #saving sparse data
                    n = n + 1
                    if n == 1: 
                        setheadingsExcelSheet(sheet, saveparameters)
                    data = len(saveparameters[0])*[0]
                    for s in range(len(data)): data[s] = locals() [saveparameters[0][s]]
                    writetoExcelSheet(sheet, n+1, saveparameters, data)
        book.save(bookname)
        book.close()
    
    sheetname = sheetnames[2]
    [f2st, f2pv] = extractSpikeTimes(tg[0:len(f2)], f2, 100) 
    [Vfst, Vfpv] = extractSpikeTimes(tg[0:len(Vfinal)], Vfinal, 100)
    [Vcst, Vcpv] = extractSpikeTimes(tg[0:len(Vcontrol)], Vcontrol, 100)
    varnames = ['f2st_i', 'f2pv_i', 'Vfst_i', 'Vfpv_i', 'Vcst_i', 'Vcpv_i']
    labels = ["Spike Time (ms)", "Original Glutamate Signal", "Spike Time (ms)", "Signal Recall", "Spike Time (ms)", "Signal Recall No Shift"]
    columns = list(range(3, 3 + len(varnames)))
    datalists = [f2st, f2pv, Vfst, Vfpv, Vcst, Vcpv]
    saveLists(datalists[0:2], 1, bookname, sheetname, [varnames[0:2],labels[0:2], columns[0:2]])
    saveLists(datalists[2:4], 1, bookname, sheetname, [varnames[2:4],labels[2:4], columns[2:4]])
    saveLists(datalists[4:6], 1, bookname, sheetname, [varnames[4:6],labels[4:6], columns[4:6]])
    
    return datalists, stable_syns, tau_dens_stable, mem_list, Vsyns, Vden_array, f2, Vfinal, Vcontrol
    

# %%  
# ****************Additional minor functions*****************************
def check_gL(num_pts, time_step, sig, decay_const, sig_const):
    tg = [0]*num_pts
    gL =[0]*num_pts
    tt = 0.                       #start time
    for j in range(num_pts):
        tt = tt + time_step
        tg[j] = tt
        gL[j] = decay_const * gL[j-1] + sig_const * sig[j]
    return tg,sig,gL

def shiftSignal(f,delayindex):
    fnew = [0]*len(f)
    for i in range (len(fnew)):
        if i >= delayindex: fnew[i] = f[i-delayindex] 
    return fnew    
    
def normalize(g):
    num_pts = len(g)
    gmax = max(g)
    gnormalized = [0]*num_pts
    for i in range(num_pts):
        gnormalized[i] = g[i]/gmax
    return gnormalized
    
def calculateDerivatives(g,t):
    num_pts = len(g)
    derivatives = [0]*(num_pts-1)
    for i in range(num_pts-1):
        derivatives[i] = g[i+1] - g[i]
    return derivatives

    
def combinedSaveIndeces(tau_d, f, f2, num_pts, t, windowlength, ranges, saveperiod_coarse, saveperiod_fine):
    delayindex = int(tau_d/t)
    f = shiftSignal(f,delayindex)
    fsavetimes = generateSaveIndeces(f, num_pts, windowlength, ranges, saveperiod_coarse, saveperiod_fine)
    f2savetimes = generateSaveIndeces(f2, num_pts, windowlength, ranges, saveperiod_coarse, saveperiod_fine)
    scombined = list(set(fsavetimes + f2savetimes))
    savetimes = sorted(scombined)
    return savetimes
    
def generateSaveIndeces(f, num_pts, windowlength, ranges, saveperiod_coarse, saveperiod_fine):
    num_pts = len(f)
    if ranges == 0: ranges = range(num_pts)
    savetimes = []
    w = windowlength
    h = int(w/2)
    savetimes = []
    for i in ranges:
        if i >= w:
            mid = i-h
            fpeak = True
            j = 0
            while fpeak and j < h:
                j = j + 1
                if f[mid-j] >= f[mid] or f[mid] <= f[mid+j]: fpeak = False
            if (fpeak or (i-w) % saveperiod_coarse == 0) and (len(savetimes) == 0 or (i-w) > savetimes[-1]):
                savetimes.append(i-w)
                if fpeak:
                    for j in range(w-1):
                        if  (i-j) % 2 == 0:
                            savetimes.append(i-j)
                    for j in range(8*w):
                        if  (i+j) < num_pts and (i+j) % saveperiod_fine == 0:
                            savetimes.append(i+j)  
    return savetimes
  
