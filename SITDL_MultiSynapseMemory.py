# -*- coding: utf-8 -*-
"""
Created on Sun Jul 18 15:38:53 2021

@author: Katya

Generates Excel data files for SITDL multi synapse runs.
Run current Initialization section, then run the desired cell section to generate Excel files.
"""
import openpyxl
import xlsxwriter
import os
import inspect
src_file_path = inspect.getfile(lambda: None)
os.path.join(src_file_path)

import math as mth
import random 
import numpy
import matplotlib.pyplot as plt  
from SITDL_Functions import saveLists, readList, generateSignals, generateRecallSpike, shiftSignal, combinedSaveIndeces,\
                                SITDL, plotSITDLOutput, SITDL_TauLoop, generateInitialTaus, \
                                SITDL_MultiSynapseMemory_Learn, SITDL_MultiSynapseMemory_SynElimination, SITDL_MultiSynapseMemory_Recall, \
                                extractSpikeTimes, normalize, calculateDerivatives
                             
#**************INITIALIZATION************************************************

# constants used for all subsequent function calls, unless overidden in section:
t=0.01                                                  # integration time step
dts=0.05                                               # tau_glu time step
num_pts=150000000 #2000000                                # number of points
recall_num_pts = int(num_pts/100)
intervals_list = [20, 14, 22, 30]                  # determines repeated sequence of inter-spike intervals
I0 = 0.01                                                 # DC current
a = 3.9                                                  # constant, determines contribution of dendritic signal to voltage
b = 0.4                                                 # constant, determines contribution of synaptic current to voltage
tau_glu_initial = 20
savetime = 100000
savetime_recall = 10
save_to_Excel = True

# %%  
#**********************************************************************************
#             SITDL MULIT-SYNAPSE MEMORY SIMULATIONS
#**********************************************************************************
num_synapses = 49
tau_den_min = 4
tau_den_max = 100
stabilization = True
thresh_factor = 1.0215

bookname = "SITDLwithP MultiSynapseMemory syn" + str(num_synapses) + ",thresh" + str(thresh_factor)+ " DATA.xlsx"
sheetnames = ["Learning Phase Final Values","Recall " + str(num_synapses) + "os, " + str(tau_den_min) + "-" + str(tau_den_max), "SpikeTimes", "SpikeTimes"]
[tau_den_list, tau_glu_list] = generateInitialTaus(num_synapses, tau_den_min, tau_den_max, tau_glu_initial, tau_glu_initial)
[f, f2]  = generateSignals(num_pts, t, intervals_list, 1., 30., 10., True, False)[1:3]    # learning signal generation (sparse V signal)

# SITDL Multi-Synapse Learning Phase
[tau_glu_list, plasticity_list, mem_list] = SITDL_MultiSynapseMemory_Learn(tau_den_list, tau_glu_list, f, f2, num_pts, t, dts, stabilization, savetime, bookname, sheetnames[0], save_to_Excel)

# Synatic Elimination    
#tau_glu_list = readList("TauGlu", bookname, sheetnames[0])
#mem_list = readList("Gmemory", bookname, sheetnames[0])
#tau_den_list = readList("TauD", bookname, sheetnames[0])
[stable_syns, tau_glu_list] = SITDL_MultiSynapseMemory_SynElimination(tau_den_list, tau_glu_list, mem_list, thresh_factor)

# SITDL Multi-Synapse Recall Phase       
f = f[0:(recall_num_pts)]       # Shortens previously generated input signals, for less memory usage
f2 = f2[0:(recall_num_pts)]
[fs, fs2] = generateRecallSpike(recall_num_pts, t, 0, 1., 30., 10.)[1:3]               # recall signal spike generation
#[stimes, stable_syns, tau_dens_stable, mem_list, Vsyns, Vden_array, f, Vfinal, Vcontrol] = SITDL_MultiSynapseMemory_Recall(tau_glu_initial, stable_syns, tau_den_list, tau_glu_list, mem_list, f, fs, f2, fs2, recall_num_pts, t, dts, savetime_recall, bookname, sheetnames, save_to_Excel)
tau_dens_stable = SITDL_MultiSynapseMemory_Recall(tau_glu_initial, stable_syns, tau_den_list, tau_glu_list, mem_list, f, fs, f2, fs2, recall_num_pts, t, dts, savetime_recall, bookname, sheetnames, save_to_Excel)[2]

# %%
#**************TauLoop data for stabilized synapses
[f, f2]  = generateSignals(num_pts, t, intervals_list, 1., 30., 10., True, False)[1:3]    # learning signal generation (sparse V signal)
tau_den_list = tau_dens_stable  #dendritic delays, based on distance between synapses
tau_glu_list = [tau_glu_initial]                                                                        #starting synaptic time lag
sheetnames_tauloop = ["Tau Glu = " + s for s in list(map(str,tau_glu_list))]
varnames = ['tt'] + len(tau_den_list)*['tau_glu']
labels = ["Time (ms)"] + list(map(str,tau_den_list))
columns = list(range(3, 4 + len(tau_den_list)))
saveparameters_loop = [varnames, labels, columns]

SITDL_TauLoop(tau_den_list, tau_glu_list, f, f2, num_pts, t, dts, I0, a, b, True, savetime, bookname, sheetnames_tauloop, saveparameters_loop)

# %%
#********** Cuts down previously simulated synapse list, runs elimination and recall again*******************************
#
starting_element = 2            #starts with tau_den = 8 ms
every_nth_element = 4           #tau_den step = 8 ms
tau_glu_list = readList("TauGlu", bookname, sheetnames[0])[starting_element::every_nth_element]
mem_list = readList("Gmemory", bookname, sheetnames[0])[starting_element::every_nth_element]
tau_den_list = readList("TauD", bookname, sheetnames[0])[starting_element::every_nth_element]

num_synapses = len(tau_glu_list) #25
tau_den_min = min(tau_den_list)  #4               
tau_den_max = max(tau_den_list)  #100
stabilization = True
thresh_factor = 1.002

bookname = "SITDLwithP MultiSynapseMemory syn" + str(num_synapses) + ",thresh" + str(thresh_factor)+ " DATA.xlsx"
sheetnames = ["Learning Phase Final Values","Recall " + str(num_synapses) + "os, " + str(tau_den_min) + "-" + str(tau_den_max), "SpikeTimes", "SpikeTimes"]
saveLists([tau_den_list, tau_glu_list, mem_list],1, bookname, sheetnames[0], [[''],["TauD","TauGlu", "Gmemory"],[3,4,5]])
[stable_syns, tau_glu_list] = SITDL_MultiSynapseMemory_SynElimination(tau_den_list, tau_glu_list, mem_list, thresh_factor)
f = f[0:(recall_num_pts)]       # Shortens previously generated input signals, for less memory usage
f2 = f2[0:(recall_num_pts)]

#tau_glu_list = readList("TauGlu", bookname, sheetnames[0])
#mem_list = readList("Gmemory", bookname, sheetnames[0])
#tau_den_list = readList("TauD", bookname, sheetnames[0])
[stable_syns, tau_glu_list] = SITDL_MultiSynapseMemory_SynElimination(tau_den_list, tau_glu_list, mem_list, thresh_factor)
tau_dens_stable = SITDL_MultiSynapseMemory_Recall(tau_glu_initial, stable_syns, tau_den_list, tau_glu_list, mem_list, f, fs, f2, fs2, recall_num_pts, t, dts, savetime_recall, bookname, sheetnames, save_to_Excel)[2]

# %%
bookname = "SITDLwithP MultiSynapseMemory SingleRun DATA.xlsx"
tau_den_list = [64]                          #dendritic delay, based on distance between synapses
tau_glu_list = [20]                                       #starting synaptic time lag
varnames = ['tt', 'tau_glu', 'f2_i', 'Vden', 'gGlu', 'gV', 'g', 'g_mismatch', 'plasticity']
labels = ["Time (ms)", "Tau Glu", "Glutamate Signal", "Dendritic Signal", "NMDAR Glutamate Gate Conductance", "NMDAR Voltage Gate Conductance", "NMDAR Conductance", "NMDAR Gate Conductance Mismatch", "Plasticity"]
columns = [3, 4, 5, 6, 7, 8, 9, 10, 11]
saveparameters = [varnames, labels, columns]
saveranges = list(range(0,30000)) + list(range(num_pts-10000,num_pts))
[f, f2]  = generateSignals(num_pts, t, intervals_list, 1., 30., 10., True, False)[1:3]       # learning signal generation (sparse V signal)
for ts in range(len(tau_den_list)):
    sheetname = "Learn. Tau Den = " + str(round(tau_den_list[ts],1))
    tau_glu_final = SITDL(tau_den_list[ts], tau_glu_list[ts], f, f2, num_pts, t, dts, I0, a, b, True, saveranges, bookname, sheetname, saveparameters)[1][-1]
del f, f2

varnames = ['tt', 'tau_glu', 'f2_i', 'Vden', 'gGlu', 'gV', 'g', 'g_mismatch', 'plasticity', 'II', 'Isyn']
labels = ["Time (ms)", "Tau Glu", "Glutamate Signal", "Dendritic Signal", "NMDAR Glutamate Gate Conductance", "NMDAR Voltage Gate Conductance", "NMDAR Conductance", "NMDAR Gate Conductance Mismatch", "Plasticity", 'AMPAR Current', 'NMDAR Current']
columns = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]
saveparameters = [varnames, labels, columns] 
saveranges = list(range(0,30000))
[fs, fs2] = generateRecallSpike(recall_num_pts, t, 0, 1., 30., 10.)[1:3]               # recall signal spike generation  
for ts in range(len(tau_den_list)):
    sheetname = "Recall. Tau Den = " + str(round(tau_den_list[ts],1))
    SITDL(0, tau_glu_final, fs, fs2, recall_num_pts, t, 0, 0., 2.0, 3.0, -1, saveranges, bookname, sheetname, saveparameters)
# %%
#*****************************************************************************************
#             CONTROL: Synaptic elimination and recall, but WITHOUT SITDL mechanisms
#*****************************************************************************************
num_synapses = 49
tau_den_min = 4
tau_den_max = 100
stabilization = False
thresh_factor = 1.0215
bookname = "woutSITDL MultiSynapseMemory syn" + str(num_synapses) + ",thresh" + str(thresh_factor)+ " DATA.xlsx"

[f, f2]  = generateSignals(num_pts, t, intervals_list, 1., 30., 10., True, False)[1:3]       # learning signal generation (sparse V signal)
[tau_den_list, tau_glu_list] = generateInitialTaus(num_synapses, tau_den_min, tau_den_max, tau_glu_initial, tau_glu_initial)
sheetnames = ["Learning Phase Final Values","Recall " + str(num_synapses) + "os, " + str(tau_den_min) + "-" + str(tau_den_max), "SpikeTimes", "SpikeTimes"]


[tau_glu_list, plasticity_list, mem_list] = SITDL_MultiSynapseMemory_Learn(tau_den_list, tau_glu_list, f, f2, num_pts, t, dts, -1, savetime, bookname, sheetnames[0], save_to_Excel)
[stable_syns, tau_glu_list] = SITDL_MultiSynapseMemory_SynElimination(tau_den_list, tau_glu_list, mem_list, thresh_factor)

f = f[0:(recall_num_pts)]
f2 = f2[0:(recall_num_pts)]
[fs, fs2] = generateRecallSpike(recall_num_pts, t, 0, 1., 30., 10.)[1:3]               # recall signal spike generation
tau_dens_stable = SITDL_MultiSynapseMemory_Recall(tau_glu_initial, stable_syns, tau_den_list, tau_glu_list, mem_list, f, fs, f2, fs2, recall_num_pts, t, dts, savetime_recall, bookname, sheetnames, save_to_Excel)[2]


#********** Cuts down previously simulated synapse list to half, runs elimination and recall again*******************************
starting_element = 2            #starts with tau_den = 8 ms
every_nth_element = 4           #tau_den step = 8 ms
tau_glu_list = readList("TauGlu", bookname, sheetnames[0])[starting_element::every_nth_element]
mem_list = readList("Gmemory", bookname, sheetnames[0])[starting_element::every_nth_element]
tau_den_list = readList("TauD", bookname, sheetnames[0])[starting_element::every_nth_element]

num_synapses = len(tau_glu_list) #12
tau_den_min = min(tau_den_list)  #8               
tau_den_max = max(tau_den_list)  #100
stabilization = True
thresh_factor = 1.002

bookname = "woutSITDL MultiSynapseMemory syn" + str(num_synapses) + ",thresh" + str(thresh_factor)+ " DATA.xlsx"
sheetnames = ["Learning Phase Final Values","Recall " + str(num_synapses) + "os, " + str(tau_den_min) + "-" + str(tau_den_max), "SpikeTimes", "SpikeTimes"]
[stable_syns, tau_glu_list] = SITDL_MultiSynapseMemory_SynElimination(tau_den_list, tau_glu_list, mem_list, thresh_factor)
tau_dens_stable = SITDL_MultiSynapseMemory_Recall(tau_glu_initial, stable_syns, tau_den_list, tau_glu_list, mem_list, f, fs, f2, fs2, recall_num_pts, t, dts, savetime_recall, bookname, sheetnames, save_to_Excel)[2]


